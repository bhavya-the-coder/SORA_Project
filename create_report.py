from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------
# Project paths
# ---------------------------------

BASE_DIR = Path(__file__).resolve().parent

DATA_FOLDER = BASE_DIR / "data"

CSV_FILE = DATA_FOLDER / "historical_sora.csv"

EXCEL_FILE = DATA_FOLDER / "SORA_Report.xlsx"


def is_file_open(filename):

    if not filename.exists():
        return False

    try:
        with open(filename, "a"):
            pass

        return False

    except PermissionError:
        return True


def create_excel_report():

    if not CSV_FILE.exists():

        print("CSV file not found.")

        return

    if is_file_open(EXCEL_FILE):

        print()
        print("SORA_Report.xlsx is open.")
        print("Please close Excel and run again.")

        return

    records = []

    # ---------------------------------
    # Read CSV
    # ---------------------------------

    with open(CSV_FILE, "r", newline="") as file:

        reader = csv.DictReader(file)

        for row in reader:
            records.append(row)

    if len(records) == 0:

        print("No data available.")

        return

    # ---------------------------------
    # Latest first
    # ---------------------------------

    records.reverse()

    # ---------------------------------
    # Create workbook
    # ---------------------------------

    wb = Workbook()

    ws = wb.active

    ws.title = "Historical SORA"

    # ---------------------------------
    # Headers
    # ---------------------------------

    headers = [
        "SORA VALUE\nDATE",
        "SORA\nPUBLICATION\nDATE",
        "SORA",
        "AGGREGATE VOLUME OF\nSORA TRANSACTIONS\n(S$ MILLIONS)",
        "HIGHEST\nTRANSACTED\nRATE",
        "LOWEST\nTRANSACTED\nRATE",
    ]

    ws.append(headers)

    # ---------------------------------
    # Header formatting
    # ---------------------------------

    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.fill = PatternFill(
            "solid",
            fgColor="DDDDDD",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # ---------------------------------
    # Header row height
    # ---------------------------------

    ws.row_dimensions[1].height = 60

    # ---------------------------------
    # Add data
    # ---------------------------------

    for row in records:

        ws.append(
            [
                row["SORA VALUE DATE"],
                row["SORA PUBLICATION DATE"],
                float(row["SORA"]),
                float(row["AGGREGATE VOLUME OF SORA TRANSACTIONS (S$ MILLIONS)"]),
                float(row["HIGHEST TRANSACTED RATE"]),
                float(row["LOWEST TRANSACTED RATE"]),
            ]
        )

    # ---------------------------------
    # Format all cells
    # ---------------------------------

    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # ---------------------------------
    # Number formats
    # ---------------------------------

    # SORA
    for cell in ws["C"][1:]:
        cell.number_format = "0.0000"

    # Aggregate volume
    for cell in ws["D"][1:]:
        cell.number_format = "0.00"

    # Highest rate
    for cell in ws["E"][1:]:
        cell.number_format = "0.0000"

    # Lowest rate
    for cell in ws["F"][1:]:
        cell.number_format = "0.0000"

    # ---------------------------------
    # Column widths
    # ---------------------------------

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13

    # ---------------------------------
    # Freeze header
    # ---------------------------------

    ws.freeze_panes = "A2"

    # ---------------------------------
    # Save Excel
    # ---------------------------------

    wb.save(EXCEL_FILE)

    print("Excel report created successfully.")
